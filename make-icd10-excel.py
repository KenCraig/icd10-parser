import re
import sys
from openpyxl import Workbook
from openpyxl.styles import Font
from peewee import *
from icd10sql import ICD10Category, ICD10SubCategory, ICD10Code, ICD10RelatedTerm


def as_text(value):
    return str(value) if value is not None else ""


if __name__ == '__main__':
    # if (len(sys.argv) > 2):
    #     inputFile = sys.argv[1]
    #     txtInput = sys.argv[2]
    # else:
    #     print("Usage: " + sys.argv[0] + " icd10cm_tabular_YYYY.xml icd10cm_codes_YYYY.txt")
    #     exit(0)

    db = SqliteDatabase('icd10codes.db')
    db.bind([ICD10Code, ICD10Category, ICD10SubCategory, ICD10RelatedTerm])
    db.connect()

    categories = ICD10Category.select()
    header = Font(size=16, bold=True)

    for category in categories:
        wb = Workbook()
        index = wb.active
        index.title = 'Index'

        subcats = ICD10SubCategory.select().where(ICD10SubCategory.category_id == category.category_id)

        for index_row, cat in enumerate(subcats, start=1):
            diags = ICD10Code.select().where(ICD10Code.subcat_id == cat.subcat_id)

            # Some categories have been deprecated but are still in the catalog
            # leave this commented out if you want blank sheets and a pure index
            # uncomment to get so sheets of 0 records, but there will be gaps in the index page
            # if not diags:
            #     continue

            ws = wb.create_sheet(cat.subcat_id)  # insert at the end (default)
            a1 = ws['A1']
            a1.value = cat.subcat
            a1.font = header

            index.cell(row=index_row, column=1, value=cat.subcat_id)  # code range
            index.cell(row=index_row, column=2, value=cat.subcat)  # subcategory
            index.cell(row=index_row, column=3, value=category.category_name)  # category

            for col, val in enumerate(['code', 'long_desc', 'morbidity', 'severity'], start=1):
                ws.cell(row=3, column=col, value=val)

            for row, diag in enumerate(diags, start=4):
                ws.cell(row=row, column=1, value=diag.diag_code)
                ws.cell(row=row, column=2, value=diag.long_desc)

        catcodes = re.findall(r'\((.*\-.*)\)', category.category_name)[0]
        filename = "".join(
            [c for c in category.category_name if c.isalpha() or c.isdigit() or c == ' ' or re.match(r'[\w\-\(\)]', c)]).rstrip()
        print(catcodes, filename)
        wb.save(f"codes-xlsx/{catcodes} {filename}.xlsx")
