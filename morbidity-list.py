import re
import csv
import sys
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font


def as_text(value):
    return str(value) if value is not None else ""


if __name__ == '__main__':
    # if (len(sys.argv) > 2):
    #     inputFile = sys.argv[1]
    #     txtInput = sys.argv[2]
    # else:
    #     print("Usage: " + sys.argv[0] + " icd10cm_tabular_YYYY.xml icd10cm_codes_YYYY.txt")
    #     exit(0)

    # icd10CSV = csv.writer(open('icd10codes.csv', 'w', encoding='utf-8'), quoting=csv.QUOTE_NONNUMERIC)
    # icd10CSV.writerow(["code", "name", "category", "subcatid", "subcat"])

    db = sqlite3.connect('icd10merge.db')

    # cursor = db.cursor()
    # result = db.execute("select distinct substr(id,1,1) as prefix from xmlsubcats")
    result = db.execute("select distinct category from xmlsubcats")
    categories = result.fetchall()
    header = Font(size=16, bold=True)

    for category in categories:
        wb = Workbook()
        index = wb.active
        index.title = 'Index'

        result = db.execute(
            "select substr(id,1,1) as prefix, id, category, subcat from xmlsubcats where category = '%s'" % category)

        subcats = result.fetchall()
        for index_row, cat in enumerate(subcats, start=1):
            result = db.execute(
                "select xmlcodes.code, long_desc from xmlcodes left join txtcodes on xmlcodes.code = txtcodes.code where subcatid = '%s'" %
                cat[1])
            diags = result.fetchall()

            # Some categories have been deprecated but are still in the catalog
            # leave this commented out if you want blank sheets and a pure index
            # uncomment to get so sheets of 0 records, but there will be gaps in the index page
            # if not diags:
            #     continue

            ws = wb.create_sheet(cat[1])  # insert at the end (default)
            a1 = ws['A1']
            a1.value = cat[3]
            a1.font = header

            index.cell(row=index_row, column=1, value=cat[1])  # code range
            index.cell(row=index_row, column=2, value=cat[3])  # subcategory
            index.cell(row=index_row, column=3, value=cat[2])  # category

            # outcsv = csv.writer(open(f"csvdump/{cat[1]}.tsv", 'w', encoding='utf-8'), dialect='excel-tab')
            outcsv = csv.writer(open(f"/dev/null", 'w', encoding='utf-8'), dialect='excel-tab')
            outcsv.writerow([cat[2]])
            outcsv.writerow(['code', 'long_desc', 'morbidity', 'severity'])

            for col, val in enumerate(['code', 'long_desc', 'morbidity', 'severity'], start=1):
                ws.cell(row=3, column=col, value=val)

            for row, diag in enumerate(diags, start=4):
                outcsv.writerow([diag[0], diag[1], None, None])
                # print("\t{0}\t{1}".format(diag[0], diag[1]))

                for col, val in enumerate(diag, start=1):   # write all diag values into cells
                    ws.cell(row=row, column=col, value=val)

        catcodes = re.findall(r'\((.*\-.*)\)', category[0])[0]
        filename = "".join(
            [c for c in category[0] if c.isalpha() or c.isdigit() or c == ' ' or re.match(r'[\w\-\(\)]', c)]).rstrip()
        print(catcodes, filename)
        wb.save(f"codes-xlsx/{catcodes} {filename}.xlsx")
